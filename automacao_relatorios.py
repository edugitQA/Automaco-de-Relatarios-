# coding: utf-8
"""
Script para automatizar a consolidação de dados e geração de relatórios
para o cliente Sousa.

Versão 17: Remove geração de relatório Word e adiciona geração de PDFs separados
           para as abas Consolidado e Dashboard Final.
"""

import pandas as pd
import openpyxl
from docx import Document
from docx.shared import Inches # Importado para definir tamanho da imagem
import os
import subprocess
import shutil
import logging
import time

# Configuração de Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s') # Corrigido: usar aspas simples

# --- Configurações --- 
BASE_DIR = "/home/klab/Downloads/edu/Automaco-de-Relatorios-"
UPLOAD_DIR = os.path.join(BASE_DIR, "upload")
FC_ORIGINAL = os.path.join(UPLOAD_DIR, "FC - Sousa.xlsx")
AV_ORIGINAL = os.path.join(UPLOAD_DIR, "AV - Sousa.xlsx")
DADOS_ORIGINAL = os.path.join(UPLOAD_DIR, "Dados - Sousa - Janeiro 2023.xlsx")
RELATORIO_TEMPLATE = os.path.join(UPLOAD_DIR, "Relatorio Tecnico - Sousa - Janeiro 2023.docx")

SENHA_FC = "fc"
SENHA_AV = "av"
SENHA_DD = "dd"

FC_DECRYPTED = os.path.join(BASE_DIR, "FC - Sousa_decrypted.xlsx")
AV_DECRYPTED = os.path.join(BASE_DIR, "AV - Sousa_decrypted.xlsx")
DADOS_DECRYPTED = os.path.join(BASE_DIR, "Dados - Sousa_decrypted_temp.xlsx")

MES_ANO = "Janeiro_2023"
MES_ABV_FC = "JAN"
MES_ABV_AV = "Jan"

try:
    parts = os.path.basename(DADOS_ORIGINAL).split(' - ')
    if len(parts) > 2:
        mes_ano_str = parts[2].split('.')[0]
        MES_ANO = mes_ano_str.replace(" ", "_")
        mes_str = mes_ano_str.split(' ')[0]
        mes_map_fc = {
            "Janeiro": "JAN", "Fevereiro": "FEV", "Março": "MAR", "Abril": "ABR",
            "Maio": "MAI", "Junho": "JUN", "Julho": "JUL", "Agosto": "AGO",
            "Setembro": "SET", "Outubro": "OUT", "Novembro": "NOV", "Dezembro": "DEZ"
        }
        mes_map_av = {
            "Janeiro": "Jan", "Fevereiro": "Fev", "Março": "Mar", "Abril": "Abr",
            "Maio": "Mai", "Junho": "Jun", "Julho": "Jul", "Agosto": "Ago",
            "Setembro": "Set", "Outubro": "Out", "Novembro": "Nov", "Dezembro": "Dez"
        }
        MES_ABV_FC = mes_map_fc.get(mes_str, "JAN")
        MES_ABV_AV = mes_map_av.get(mes_str, "Jan")
        logging.info(f"Mês/Ano detectado: {MES_ANO}. Aba FC: FC - {MES_ABV_FC}. Aba AV: {MES_ABV_AV}")
except Exception as e:
    logging.warning(f"Não foi possível extrair Mês/Ano do nome do arquivo Dados. Usando defaults: {MES_ANO}, FC - {MES_ABV_FC}, {MES_ABV_AV}. Erro: {e}")

RELATORIO_FINAL = os.path.join(BASE_DIR, f"Relatorio_Tecnico_Sousa_{MES_ANO}_Automatizado.docx")
DADOS_FINAL = os.path.join(BASE_DIR, f"Dados_Sousa_{MES_ANO}_Processado.xlsx")

# --- Funções Auxiliares ---

def decrypt_file(encrypted_file, decrypted_file, password):
    if not os.path.exists(encrypted_file):
        logging.error(f"Arquivo criptografado não encontrado: {encrypted_file}")
        return False
    try:
        os.makedirs(os.path.dirname(decrypted_file), exist_ok=True)
        if os.path.exists(decrypted_file):
            os.remove(decrypted_file)
        # Só descriptografa se for .xls (OLE)
        if encrypted_file.lower().endswith('.xls'):
            logging.info(f"Descriptografando {os.path.basename(encrypted_file)}...")
            cmd = ["msoffcrypto-tool", encrypted_file, decrypted_file, "-p", password]
            result = subprocess.run(cmd, capture_output=True, text=True, check=True)
            logging.info(f"Arquivo descriptografado salvo como: {decrypted_file}")
        else:
            shutil.copy(encrypted_file, decrypted_file)
            logging.info(f"Arquivo não é .xls protegido. Copiado para: {decrypted_file}")
        return True
    except subprocess.CalledProcessError as e:
        logging.error(f"Erro ao descriptografar {os.path.basename(encrypted_file)}. Stderr: {e.stderr}")
        if os.path.exists(decrypted_file): os.remove(decrypted_file)
        return False
    except Exception as e:
        logging.error(f"Erro inesperado ao processar {os.path.basename(encrypted_file)}: {e}")
        if os.path.exists(decrypted_file): os.remove(decrypted_file)
        return False

def cleanup_temp_files(files):
    for file in files:
        if file and os.path.exists(file):
            try:
                os.remove(file)
                logging.info(f"Arquivo temporário removido: {file}")
            except Exception as e:
                logging.warning(f"Não foi possível remover o arquivo temporário {file}: {e}")

# --- Funções Principais --- 

def extrair_dados_fc_detalhado(arquivo_fc_decrypted):
    logging.info(f"Extraindo dados detalhados de {os.path.basename(arquivo_fc_decrypted)}...")
    dados_detalhados = []
    try:
        excel_file = pd.ExcelFile(arquivo_fc_decrypted)
        sheet_names = excel_file.sheet_names
        logging.info(f"Abas encontradas em FC: {sheet_names}")
        aba_fc_mensal = f"FC - {MES_ABV_FC}"
        if aba_fc_mensal not in sheet_names:
            logging.error(f"Aba mensal '{aba_fc_mensal}' não encontrada no arquivo FC. Abortando extração detalhada.")
            return None
        logging.info(f"Lendo dados detalhados da aba FC: {aba_fc_mensal}")
        workbook = openpyxl.load_workbook(arquivo_fc_decrypted, data_only=True)
        sheet = workbook[aba_fc_mensal]
        linha_inicial = 7
        logging.info(f"Iniciando leitura a partir da linha {linha_inicial}")
        for row_idx in range(linha_inicial, sheet.max_row + 1):
            cod = sheet.cell(row=row_idx, column=2).value
            conta = sheet.cell(row=row_idx, column=3).value
            realizado = sheet.cell(row=row_idx, column=5).value
            if cod is None and conta is None and realizado is None:
                logging.info(f"Linha {row_idx} vazia encontrada, parando a extração.")
                break
            if isinstance(conta, str) and conta.strip().upper() in ["RECEITAS", "SAÍDAS"]:
                logging.info(f"Ignorando linha {row_idx} (cabeçalho de seção: {conta})")
                continue
            if cod is not None or conta is not None or realizado is not None:
                dados_detalhados.append({"cod": cod, "conta": conta, "realizado": realizado})
                if (row_idx - linha_inicial + 1) % 5 == 0:
                     logging.info(f"Extraída linha {row_idx}: Cod={cod}, Conta={conta}, Realizado={realizado}")
            else:
                 logging.info(f"Ignorando linha {row_idx} (vazia ou sem dados relevantes)")
        logging.info(f"Extração detalhada da aba {aba_fc_mensal} concluída. {len(dados_detalhados)} linhas extraídas.")
        return dados_detalhados
    except Exception as e:
        logging.error(f"Erro ao extrair dados detalhados de FC: {e}")
        return None

def extrair_dados_av(arquivo_av_decrypted):
    logging.info(f"Extraindo dados de vendas de {os.path.basename(arquivo_av_decrypted)}...")
    try:
        excel_file = pd.ExcelFile(arquivo_av_decrypted)
        sheet_names = excel_file.sheet_names
        logging.info(f"Abas encontradas em AV: {sheet_names}")
        aba_av_mensal = MES_ABV_AV
        if aba_av_mensal not in sheet_names:
            logging.error(f"Aba mensal '{aba_av_mensal}' não encontrada no arquivo AV. Abortando extração de vendas.")
            return None
        logging.info(f"Lendo dados da aba AV: {aba_av_mensal}")
        df_av = pd.read_excel(arquivo_av_decrypted, sheet_name=aba_av_mensal)
        coluna_g_idx = 6
        total_vendas = None
        if df_av.shape[1] > coluna_g_idx:
            coluna_g = df_av.iloc[:, coluna_g_idx].dropna()
            if not coluna_g.empty:
                for valor in reversed(coluna_g.values):
                    if isinstance(valor, (int, float)):
                        total_vendas = valor
                        logging.info(f"Total de vendas encontrado no final da coluna G da aba {aba_av_mensal}: {total_vendas}")
                        break
                if total_vendas is None: logging.warning(f"Nenhum valor numérico encontrado no final da coluna G da aba {aba_av_mensal}.")
            else: logging.warning(f"Coluna G da aba {aba_av_mensal} está vazia ou contém apenas NaNs.")
        else: logging.warning(f"Coluna G (índice {coluna_g_idx}) não encontrada na aba {aba_av_mensal}.")
        if total_vendas is not None: return {"total_vendas": total_vendas}
        else: logging.error(f"Não foi possível extrair o total de vendas da aba {aba_av_mensal}."); return None
    except Exception as e:
        logging.error(f"Erro ao extrair dados de AV: {e}")
        return None

def consolidar_dados(arquivo_dados_template, arquivo_dados_processado, dados_fc_detalhados, dados_av):
    logging.info(f"Consolidando dados em {os.path.basename(arquivo_dados_processado)}...")
    try:
        shutil.copy(arquivo_dados_template, arquivo_dados_processado)
        workbook = openpyxl.load_workbook(arquivo_dados_processado, keep_vba=True)
        sheet_names_dados = workbook.sheetnames
        logging.info(f"Abas encontradas em Dados: {sheet_names_dados}")
        aba_destino_fc_nome = "R.T - Fluxo de Caixa"
        if aba_destino_fc_nome not in sheet_names_dados:
            logging.error(f"Aba de destino '{aba_destino_fc_nome}' não encontrada na planilha Dados. Abortando consolidação FC.")
        else:
            logging.info(f"Usando aba de destino para FC: {aba_destino_fc_nome}")
            sheet_destino_fc = workbook[aba_destino_fc_nome]
            if dados_fc_detalhados:
                linha_escrita_fc = 8
                logging.info(f"Iniciando escrita dos dados FC na aba {aba_destino_fc_nome} a partir da linha {linha_escrita_fc}")
                for i, linha_dado in enumerate(dados_fc_detalhados):
                    linha_atual = linha_escrita_fc + i
                    try:
                        sheet_destino_fc.cell(row=linha_atual, column=2).value = linha_dado.get("cod")
                        sheet_destino_fc.cell(row=linha_atual, column=4).value = linha_dado.get("conta")
                        sheet_destino_fc.cell(row=linha_atual, column=6).value = linha_dado.get("realizado")
                        if i % 10 == 0: logging.info(f"Linha {linha_atual} escrita na aba {aba_destino_fc_nome}.")
                    except Exception as e_write_fc: logging.error(f"Erro ao escrever linha {linha_atual} na aba {aba_destino_fc_nome}: {e_write_fc}")
                logging.info(f"Escrita dos dados FC detalhados concluída.")
            else: logging.warning("Dados detalhados de FC não disponíveis para escrita.")
        aba_destino_vendas_nome = "Vendas 12 M."
        if aba_destino_vendas_nome not in sheet_names_dados:
             logging.error(f"Aba de destino '{aba_destino_vendas_nome}' não encontrada na planilha Dados. Abortando consolidação Vendas.")
        else:
            logging.info(f"Usando aba de destino para Vendas: {aba_destino_vendas_nome}")
            sheet_destino_vendas = workbook[aba_destino_vendas_nome]
            if dados_av and "total_vendas" in dados_av:
                total_vendas = dados_av["total_vendas"]
                linha_escrita_vendas = 5
                mes_col_map = {"Jan": 4, "Fev": 5, "Mar": 6, "Abr": 7, "Mai": 8, "Jun": 9, "Jul": 10, "Ago": 11, "Set": 12, "Out": 13, "Nov": 14, "Dez": 15}
                coluna_escrita_vendas = mes_col_map.get(MES_ABV_AV)
                if coluna_escrita_vendas:
                    try:
                        sheet_destino_vendas.cell(row=linha_escrita_vendas, column=coluna_escrita_vendas).value = total_vendas
                        logging.info(f"Total de Vendas ({total_vendas}) escrito na aba {aba_destino_vendas_nome}, linha {linha_escrita_vendas}, coluna {coluna_escrita_vendas} ({MES_ABV_AV}).")
                    except Exception as e_write_av: logging.error(f"Erro ao escrever total de vendas na aba {aba_destino_vendas_nome}: {e_write_av}")
                else: logging.error(f"Não foi possível encontrar a coluna correspondente ao mês '{MES_ABV_AV}' na aba {aba_destino_vendas_nome}.")
            else: logging.warning("Dados de AV (total_vendas) não disponíveis para escrita.")
        workbook.save(arquivo_dados_processado)
        logging.info(f"Planilha de Dados consolidada salva como: {arquivo_dados_processado}")
        return True
    except Exception as e:
        logging.error(f"Erro ao consolidar dados na planilha Dados: {e}")
        return False

# --- Fluxo Principal ---
def main():
    logging.info("--- Iniciando processo de automação de relatórios (atualização e PDFs das abas via PNG) ---")
    temp_files_to_clean = []
    process_success = True
    try:
        success_fc = decrypt_file(FC_ORIGINAL, FC_DECRYPTED, SENHA_FC)
        if success_fc: temp_files_to_clean.append(FC_DECRYPTED)
        success_av = decrypt_file(AV_ORIGINAL, AV_DECRYPTED, SENHA_AV)
        if success_av: temp_files_to_clean.append(AV_DECRYPTED)
        success_dd = decrypt_file(DADOS_ORIGINAL, DADOS_DECRYPTED, SENHA_DD)
        if success_dd: temp_files_to_clean.append(DADOS_DECRYPTED)
        if not (success_fc and success_av and success_dd):
            logging.error("Falha na descriptografia de um ou mais arquivos. Abortando.")
            process_success = False; return
        dados_fc_detalhados = extrair_dados_fc_detalhado(FC_DECRYPTED)
        dados_av = extrair_dados_av(AV_DECRYPTED)
        if dados_fc_detalhados is None:
            logging.error("Falha na extração de dados detalhados de FC. Abortando.")
            process_success = False; return
        if dados_av is None:
            logging.warning("Falha na extração de dados de AV. Continuando sem dados de AV.")
            dados_av = {}
        success_consolidacao = consolidar_dados(DADOS_DECRYPTED, DADOS_FINAL, dados_fc_detalhados, dados_av)
        if not success_consolidacao:
            logging.error("Falha na consolidação dos dados. Abortando.")
            process_success = False; return
        logging.info("Planilha de Dados consolidada. Geração de PDFs das abas será feita pelo main.py.")
    finally:
        cleanup_temp_files(temp_files_to_clean)
        logging.info("Limpeza de arquivos temporários descriptografados concluída.")
        logging.info("--- Processo de automação concluído ---")
        if process_success:
            logging.info(f"Arquivo final gerado (verificar conteúdo): {DADOS_FINAL}")
        else:
            logging.warning("Processo concluído com erros. Verifique os logs.")

if __name__ == "__main__":
    main()


